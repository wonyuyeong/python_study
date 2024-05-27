from openpyxl import load_workbook

# data_only  -> 값만 보이게 하기.
wb = load_workbook("day04/data/excel01.xlsx", data_only=True)
w_list = wb.sheetnames
print(w_list)

# 타이틀을 이용해서 sheet 얻기
ws = wb["타이틀연습"]
print(ws["A1"].value)
print(ws["A5"].value)       # 데이터가 없으면 None
print("-" * 50)

# 범위 지정
get_cell = ws["A1" : "C4"]
for i in get_cell :
    for j in i :
        print(j.value, end=' ')
    print()
print("-" * 50)

get_cell = ws["A1" : "C4"]
for i in get_cell :
    for j in i :
        if(j.value == None) : continue
        print(j.value, end=' ')
    print()
print("-" * 50)

# 데이터를 리스트에 담기
lis = []
get_cell = ws["A1" : "C6"]
for i in get_cell :
    for j in i :
        if(j.value == None) : continue
        lis.append(j.value)
print("-" * 50)

print(lis)