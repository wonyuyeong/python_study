# 엑셀 구조

# workbook - 엑셀파일, worksheet - 엑셀시트, cell - 시트안에 존재하는 셀 하나
# 엑셀 관련 모듈
# 1) xlwt 2) OpenPyXL 3)pandas(빅데이터)...

# openpyxl를 설치 : 터미널에서 pip install openpyxl

# from 모듈 import 이름
from openpyxl import Workbook

# 파일만들기
f_path = 'day04/data/excel01.xlsx'
# Workbook() 함수를 통해서  workbook 객체 생성
# 생성시 기본적으로 sheet 하나가 생성된다.
wb = Workbook()
ws = wb.active

Sheet2 = wb.create_sheet("두번째 시트")     # 추가
Sheet3 = wb.create_sheet("세번재 시트", 1)  # 삽입

Sheet2.title = "성적관련정보"
ws.title = "타이틀연습"

# A1 cell에 데이터 쓰기  
ws['A1'] = "Hello World"
ws.cell(row=2, column=2,value="방가방가")
ws.cell(row=3, column=3) .value="하이`1~~~"
ws.cell(row=4, column=1) .value="데이터 입력 연습"

# 범위 지정해서 두번째 셀에 데이터 넣기
for i in range(5) :
    Sheet2.cell(row=1, column=i+1).value="Hi~~~"
    
# 리스트 한번에 넣기
Sheet3.append([1,2,3,4,5,6,7,8,9])

# 시트 타이틀을 이용해서 데이터 입력
ws3 = wb["세번재 시트"]
ws3['C1'] = "안녕하세요"

# 저장하고 파일 닫기
wb.save(f_path)
wb.close()
